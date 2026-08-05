from pathlib import Path
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import patch
from zipfile import ZIP_DEFLATED, ZipFile

import pandas as pd
from openpyxl import Workbook
from pydantic import ValidationError

from app.core.config import Settings, settings
from app.services.transformacion_excel_security_service import (
    TransformacionExcelSecurityError,
    calculate_file_sha256,
    calculate_transformacion_config_checksum,
    resolve_storage_path_safely,
    validate_dataframe_dimensions,
    validate_source_file_security,
    validate_xlsx_container,
)


class TransformacionExcelSecurityTests(unittest.TestCase):
    def test_safe_relative_storage_path_is_resolved_inside_root(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory) / "storage"
            expected = root / "originals" / "source.csv"
            self.assertEqual(
                resolve_storage_path_safely("originals/source.csv", root),
                expected.resolve(),
            )

    def test_legacy_storage_prefixed_path_is_supported(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory) / "storage"
            expected = root / "originals" / "source.csv"
            self.assertEqual(
                resolve_storage_path_safely("storage/originals/source.csv", root),
                expected.resolve(),
            )

    def test_storage_path_escape_is_rejected(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory) / "storage"
            with self.assertRaisesRegex(
                TransformacionExcelSecurityError,
                "UNSAFE_STORAGE_PATH",
            ):
                resolve_storage_path_safely("../secret.csv", root)

    def test_external_absolute_path_is_rejected(self) -> None:
        with TemporaryDirectory() as directory, TemporaryDirectory() as external:
            root = Path(directory) / "storage"
            with self.assertRaisesRegex(
                TransformacionExcelSecurityError,
                "UNSAFE_STORAGE_PATH",
            ):
                resolve_storage_path_safely(Path(external) / "source.csv", root)

    def test_symlink_escape_is_rejected_when_supported(self) -> None:
        with TemporaryDirectory() as directory:
            base = Path(directory)
            root = base / "storage"
            root.mkdir()
            external = base / "external"
            external.mkdir()
            link = root / "linked"
            try:
                link.symlink_to(external, target_is_directory=True)
            except OSError as exc:
                self.skipTest(f"Symlinks no disponibles en este entorno: {exc}")
            with self.assertRaisesRegex(
                TransformacionExcelSecurityError,
                "UNSAFE_STORAGE_PATH",
            ):
                resolve_storage_path_safely("linked/source.csv", root)

    def test_file_checksum_uses_physical_content(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "source.csv"
            path.write_bytes(b"a,b\n1,2\n")
            first = calculate_file_sha256(path)
            path.write_bytes(b"a,b\n1,3\n")
            self.assertNotEqual(first, calculate_file_sha256(path))

    def test_config_checksum_is_deterministic_for_dictionary_order(self) -> None:
        first = {"source": {"archivo_id": 1, "header_row": 1}, "rows": {}}
        second = {"rows": {}, "source": {"header_row": 1, "archivo_id": 1}}
        self.assertEqual(
            calculate_transformacion_config_checksum(first),
            calculate_transformacion_config_checksum(second),
        )

    def test_dataframe_row_limit_is_413_and_does_not_modify_source(self) -> None:
        dataframe = pd.DataFrame({"value": [1, 2]})
        snapshot = dataframe.copy(deep=True)
        with patch.object(settings, "transformacion_excel_max_rows", 1):
            with self.assertRaises(TransformacionExcelSecurityError) as context:
                validate_dataframe_dimensions(dataframe)
        self.assertEqual(context.exception.status_code, 413)
        self.assertIn("SOURCE_ROW_LIMIT_EXCEEDED", str(context.exception))
        pd.testing.assert_frame_equal(dataframe, snapshot)

    def test_dataframe_column_limit_is_413(self) -> None:
        dataframe = pd.DataFrame({"a": [1], "b": [2]})
        with patch.object(settings, "transformacion_excel_max_columns", 1):
            with self.assertRaisesRegex(
                TransformacionExcelSecurityError,
                "SOURCE_COLUMN_LIMIT_EXCEEDED",
            ):
                validate_dataframe_dimensions(dataframe)

    def test_file_size_limit_uses_filesystem_size(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "source.csv"
            path.write_bytes(b"x" * (1024 * 1024 + 1))
            with patch.object(settings, "transformacion_excel_max_file_size_mb", 1):
                with self.assertRaises(TransformacionExcelSecurityError) as context:
                    validate_source_file_security(path, ".csv")
            self.assertEqual(context.exception.status_code, 413)
            self.assertIn("SOURCE_FILE_TOO_LARGE", str(context.exception))

    def test_normal_xlsx_is_accepted(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "normal.xlsx"
            workbook = Workbook()
            workbook.active["A1"] = "value"
            workbook.save(path)
            validate_xlsx_container(path)

    def test_corrupt_xlsx_is_rejected(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "corrupt.xlsx"
            path.write_bytes(b"not-a-zip")
            with self.assertRaisesRegex(
                TransformacionExcelSecurityError,
                "INVALID_XLSX_CONTAINER",
            ):
                validate_xlsx_container(path)

    def test_unsafe_xlsx_entry_is_rejected(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "unsafe.xlsx"
            with ZipFile(path, "w") as archive:
                archive.writestr("../escape.txt", "unsafe")
            with self.assertRaisesRegex(
                TransformacionExcelSecurityError,
                "UNSAFE_XLSX_ENTRY",
            ):
                validate_xlsx_container(path)

    def test_excessive_xlsx_compression_ratio_is_rejected(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "ratio.xlsx"
            workbook_xml = (
                '<workbook xmlns="http://schemas.openxmlformats.org/'
                'spreadsheetml/2006/main"><sheets/></workbook>'
            )
            with ZipFile(path, "w", compression=ZIP_DEFLATED) as archive:
                archive.writestr("xl/workbook.xml", workbook_xml)
                archive.writestr("xl/worksheets/sheet1.xml", b"0" * 1_100_000)
            with patch.object(
                settings,
                "transformacion_excel_max_xlsx_compression_ratio",
                2,
            ):
                with self.assertRaisesRegex(
                    TransformacionExcelSecurityError,
                    "XLSX_COMPRESSION_RATIO_LIMIT",
                ):
                    validate_xlsx_container(path)

    def test_sheet_limit_is_enforced(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "sheets.xlsx"
            workbook = Workbook()
            workbook.create_sheet("Second")
            workbook.save(path)
            with patch.object(settings, "transformacion_excel_max_sheets", 1):
                with self.assertRaisesRegex(
                    TransformacionExcelSecurityError,
                    "WORKBOOK_SHEET_LIMIT",
                ):
                    validate_xlsx_container(path)

    def test_limit_settings_reject_non_positive_values(self) -> None:
        with self.assertRaises(ValidationError):
            Settings(transformacion_excel_max_rows=0)


if __name__ == "__main__":
    unittest.main()
