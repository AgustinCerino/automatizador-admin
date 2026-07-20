from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import load_workbook
import pandas as pd
from pandas.testing import assert_frame_equal

from app.schemas.transformacion_excel import TransformacionExcelConfig
from app.services.transformacion_excel_xlsx_writer import (
    MAX_COLUMN_WIDTH,
    TransformacionExcelXlsxWriterError,
    build_output_path,
    write_transformacion_xlsx,
)


def build_config(
    output_columns: list[dict],
    *,
    sheet_name: str = "Resultado",
    freeze_header: bool = True,
    auto_filter: bool = True,
    auto_width: bool = True,
) -> TransformacionExcelConfig:
    return TransformacionExcelConfig.model_validate(
        {
            "source": {"archivo_id": 1},
            "output_columns": output_columns,
            "output": {
                "file_name": "resultado.xlsx",
                "sheet_name": sheet_name,
                "freeze_header": freeze_header,
                "auto_filter": auto_filter,
                "auto_width": auto_width,
            },
        },
    )


def source_transform(
    output_column: str,
    output_type: str,
    *,
    position: int,
    decimal_places: int | None = None,
    date_format: str | None = None,
) -> dict:
    transform = {
        "operation": "SOURCE",
        "position": position,
        "output_column": output_column,
        "source_column": output_column,
        "output_type": output_type,
    }
    if decimal_places is not None:
        transform["decimal_places"] = decimal_places
    if date_format is not None:
        transform["date_format"] = date_format
    return transform


class TransformacionExcelXlsxWriterTests(unittest.TestCase):
    def test_generates_valid_xlsx_with_configured_sheet(self) -> None:
        dataframe = pd.DataFrame([{"Cliente": "Ana"}])
        config = build_config(
            [source_transform("Cliente", "text", position=1)],
            sheet_name="Ventas",
        )

        with tempfile.TemporaryDirectory() as directory:
            output_path = Path(directory) / "resultado.xlsx"
            write_transformacion_xlsx(dataframe, config, output_path)
            workbook = load_workbook(output_path)

            self.assertGreater(output_path.stat().st_size, 0)
            self.assertEqual(workbook.sheetnames, ["Ventas"])

    def test_preserves_exact_column_order_without_pandas_index(self) -> None:
        dataframe = pd.DataFrame(
            [{"Segundo": "B", "Primero": "A"}],
            columns=["Primero", "Segundo"],
        )
        config = build_config(
            [
                source_transform("Primero", "text", position=1),
                source_transform("Segundo", "text", position=2),
            ],
        )

        with tempfile.TemporaryDirectory() as directory:
            output_path = Path(directory) / "resultado.xlsx"
            write_transformacion_xlsx(dataframe, config, output_path)
            worksheet = load_workbook(output_path).active

            self.assertEqual(
                [worksheet.cell(1, index).value for index in (1, 2)],
                ["Primero", "Segundo"],
            )
            self.assertEqual(worksheet.max_column, 2)
            self.assertEqual(worksheet["A2"].value, "A")

    def test_empty_dataframe_writes_headers_and_valid_workbook(self) -> None:
        dataframe = pd.DataFrame(columns=["A", "B"])
        config = build_config(
            [
                source_transform("A", "text", position=1),
                source_transform("B", "integer", position=2),
            ],
        )

        with tempfile.TemporaryDirectory() as directory:
            output_path = Path(directory) / "resultado.xlsx"
            write_transformacion_xlsx(dataframe, config, output_path)
            worksheet = load_workbook(output_path).active

            self.assertEqual(worksheet.max_row, 1)
            self.assertEqual(
                [worksheet["A1"].value, worksheet["B1"].value],
                ["A", "B"],
            )

    def test_freeze_header_uses_a2(self) -> None:
        config = self._basic_config(freeze_header=True)
        with tempfile.TemporaryDirectory() as directory:
            output_path = Path(directory) / "resultado.xlsx"
            write_transformacion_xlsx(
                pd.DataFrame([{"A": "x"}]),
                config,
                output_path,
            )
            self.assertEqual(
                load_workbook(output_path).active.freeze_panes,
                "A2",
            )

    def test_freeze_header_false_does_not_freeze(self) -> None:
        config = self._basic_config(freeze_header=False)
        with tempfile.TemporaryDirectory() as directory:
            output_path = Path(directory) / "resultado.xlsx"
            write_transformacion_xlsx(
                pd.DataFrame([{"A": "x"}]),
                config,
                output_path,
            )
            self.assertIsNone(
                load_workbook(output_path).active.freeze_panes,
            )

    def test_auto_filter_covers_used_range(self) -> None:
        config = self._basic_config(auto_filter=True)
        with tempfile.TemporaryDirectory() as directory:
            output_path = Path(directory) / "resultado.xlsx"
            write_transformacion_xlsx(
                pd.DataFrame([{"A": "x"}]),
                config,
                output_path,
            )
            self.assertEqual(
                load_workbook(output_path).active.auto_filter.ref,
                "A1:A2",
            )

    def test_auto_width_applies_minimum_and_maximum(self) -> None:
        dataframe = pd.DataFrame(
            [{"Corta": "x", "Larga": "z" * 100}],
        )
        config = build_config(
            [
                source_transform("Corta", "text", position=1),
                source_transform("Larga", "text", position=2),
            ],
        )

        with tempfile.TemporaryDirectory() as directory:
            output_path = Path(directory) / "resultado.xlsx"
            write_transformacion_xlsx(dataframe, config, output_path)
            worksheet = load_workbook(output_path).active

            self.assertGreaterEqual(
                worksheet.column_dimensions["A"].width,
                10,
            )
            self.assertEqual(
                worksheet.column_dimensions["B"].width,
                MAX_COLUMN_WIDTH,
            )

    def test_headers_are_bold(self) -> None:
        config = self._basic_config()
        with tempfile.TemporaryDirectory() as directory:
            output_path = Path(directory) / "resultado.xlsx"
            write_transformacion_xlsx(
                pd.DataFrame([{"A": "x"}]),
                config,
                output_path,
            )
            self.assertTrue(load_workbook(output_path).active["A1"].font.bold)

    def test_dates_are_real_excel_dates_with_configured_format(self) -> None:
        dataframe = pd.DataFrame(
            [{"Fecha": date(2026, 7, 20)}],
        )
        config = build_config(
            [
                source_transform(
                    "Fecha",
                    "date",
                    position=1,
                    date_format="%d/%m/%Y",
                ),
            ],
        )

        with tempfile.TemporaryDirectory() as directory:
            output_path = Path(directory) / "resultado.xlsx"
            write_transformacion_xlsx(dataframe, config, output_path)
            cell = load_workbook(output_path).active["A2"]

            self.assertIsInstance(cell.value, datetime)
            self.assertEqual(cell.number_format, "dd/mm/yyyy")

    def test_decimal_uses_configured_decimal_places(self) -> None:
        dataframe = pd.DataFrame([{"Importe": Decimal("1234.50")}])
        config = build_config(
            [
                source_transform(
                    "Importe",
                    "decimal",
                    position=1,
                    decimal_places=2,
                ),
            ],
        )

        with tempfile.TemporaryDirectory() as directory:
            output_path = Path(directory) / "resultado.xlsx"
            write_transformacion_xlsx(dataframe, config, output_path)
            cell = load_workbook(output_path).active["A2"]

            self.assertEqual(cell.value, 1234.5)
            self.assertEqual(cell.number_format, "#,##0.00")

    def test_integer_uses_integer_format(self) -> None:
        dataframe = pd.DataFrame([{"Cantidad": 7}])
        config = build_config(
            [source_transform("Cantidad", "integer", position=1)],
        )

        with tempfile.TemporaryDirectory() as directory:
            output_path = Path(directory) / "resultado.xlsx"
            write_transformacion_xlsx(dataframe, config, output_path)
            cell = load_workbook(output_path).active["A2"]

            self.assertEqual(cell.value, 7)
            self.assertEqual(cell.number_format, "0")

    def test_formula_injection_is_neutralized_only_for_text(self) -> None:
        dataframe = pd.DataFrame(
            [
                {"Texto": "=1+1", "Numero": -2},
                {"Texto": "  @SUM(A1)", "Numero": 3},
            ],
        )
        config = build_config(
            [
                source_transform("Texto", "text", position=1),
                source_transform("Numero", "integer", position=2),
            ],
        )

        with tempfile.TemporaryDirectory() as directory:
            output_path = Path(directory) / "resultado.xlsx"
            write_transformacion_xlsx(dataframe, config, output_path)
            worksheet = load_workbook(output_path, data_only=False).active

            self.assertEqual(worksheet["A2"].value, "'=1+1")
            self.assertEqual(worksheet["A3"].value, "'  @SUM(A1)")
            self.assertEqual(worksheet["A2"].data_type, "s")
            self.assertEqual(worksheet["B2"].value, -2)

    def test_writer_does_not_modify_dataframe(self) -> None:
        dataframe = pd.DataFrame([{"A": "=1+1"}])
        original = dataframe.copy(deep=True)
        config = self._basic_config()

        with tempfile.TemporaryDirectory() as directory:
            write_transformacion_xlsx(
                dataframe,
                config,
                Path(directory) / "resultado.xlsx",
            )
        assert_frame_equal(dataframe, original)

    def test_atomic_write_leaves_nonempty_final_without_temporary(self) -> None:
        config = self._basic_config()
        with tempfile.TemporaryDirectory() as directory:
            output_path = Path(directory) / "resultado.xlsx"
            write_transformacion_xlsx(
                pd.DataFrame([{"A": "x"}]),
                config,
                output_path,
            )

            self.assertTrue(output_path.is_file())
            self.assertGreater(output_path.stat().st_size, 0)
            self.assertEqual(
                list(Path(directory).glob("*.tmp.xlsx")),
                [],
            )

    def test_invalid_sheet_name_is_controlled_and_writes_nothing(self) -> None:
        config = self._basic_config(sheet_name="Invalida/Hoja")
        with tempfile.TemporaryDirectory() as directory:
            output_path = Path(directory) / "resultado.xlsx"
            with self.assertRaises(TransformacionExcelXlsxWriterError):
                write_transformacion_xlsx(
                    pd.DataFrame([{"A": "x"}]),
                    config,
                    output_path,
                )
            self.assertFalse(output_path.exists())

    def test_temporary_file_is_removed_when_workbook_write_fails(self) -> None:
        config = self._basic_config()
        with tempfile.TemporaryDirectory() as directory:
            output_path = Path(directory) / "resultado.xlsx"
            with patch(
                "app.services.transformacion_excel_xlsx_writer.write_workbook",
                side_effect=RuntimeError("fallo simulado"),
            ):
                with self.assertRaises(TransformacionExcelXlsxWriterError):
                    write_transformacion_xlsx(
                        pd.DataFrame([{"A": "x"}]),
                        config,
                        output_path,
                    )

            self.assertFalse(output_path.exists())
            self.assertEqual(
                list(Path(directory).glob("*.tmp.xlsx")),
                [],
            )

    def test_path_traversal_file_name_is_rejected_before_write(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            with self.assertRaises(TransformacionExcelXlsxWriterError):
                build_output_path(
                    Path(directory),
                    "../fuera.xlsx",
                )
            self.assertFalse((Path(directory).parent / "fuera.xlsx").exists())

    def _basic_config(
        self,
        *,
        sheet_name: str = "Resultado",
        freeze_header: bool = True,
        auto_filter: bool = True,
    ) -> TransformacionExcelConfig:
        return build_config(
            [source_transform("A", "text", position=1)],
            sheet_name=sheet_name,
            freeze_header=freeze_header,
            auto_filter=auto_filter,
        )


if __name__ == "__main__":
    unittest.main()
