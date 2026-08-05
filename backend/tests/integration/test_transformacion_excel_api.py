from __future__ import annotations

from copy import deepcopy
from io import BytesIO
from pathlib import Path
import unittest
from unittest.mock import patch
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select

from app.core.config import settings
from app.core.security import get_password_hash
from app.database.session import get_db
from app.main import app
from app.models import Archivo, Cliente, EjecucionProceso, Proceso, Usuario
from app.services.transformacion_excel_generation_service import OUTPUT_FILE_TYPE
from app.services.transformacion_excel_security_service import (
    resolve_storage_path_safely,
)

from .conftest import IsolatedDatabaseAndStorage, create_test_engine


PASSWORD = "integration-secret"


def workbook_bytes(rows: list[list[object]] | None = None) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Ventas"
    worksheet.append(["Cliente", "Cantidad", "Precio", "Categoria"])
    for row in rows or [["Ana", 2, 10, "A"], ["Beto", 3, 5, "B"]]:
        worksheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def full_config(archivo_id: int, file_name: str = "resultado.xlsx") -> dict:
    return {
        "source": {
            "archivo_id": archivo_id,
            "sheet_name": "Ventas",
            "header_row": 1,
        },
        "output_columns": [
            {
                "operation": "SOURCE",
                "position": 1,
                "output_column": "Cliente",
                "source_column": "Cliente",
                "output_type": "text",
            },
            {
                "operation": "CONSTANT",
                "position": 2,
                "output_column": "Empresa",
                "value": "ACME",
                "output_type": "text",
            },
            {
                "operation": "CONCAT",
                "position": 3,
                "output_column": "Etiqueta",
                "parts": [
                    {"type": "SOURCE", "value": "Cliente"},
                    {"type": "LITERAL", "value": "-venta"},
                ],
                "output_type": "text",
            },
            {
                "operation": "ARITHMETIC",
                "position": 4,
                "output_column": "Total",
                "operator": "MULTIPLY",
                "left_operand": {"type": "SOURCE", "value": "Cantidad"},
                "right_operand": {"type": "SOURCE", "value": "Precio"},
                "output_type": "decimal",
                "decimal_places": 2,
            },
            {
                "operation": "VALUE_MAP",
                "position": 5,
                "output_column": "Categoria normalizada",
                "source_column": "Categoria",
                "mapping": {"A": "Alta", "B": "Baja"},
                "unmapped_policy": "ERROR",
                "output_type": "text",
            },
        ],
        "rows": {
            "filters": [],
            "remove_duplicates": {"enabled": False},
            "sort_by": [{"output_column": "Cliente", "direction": "ASC"}],
        },
        "output": {
            "file_name": file_name,
            "sheet_name": "Resultado",
            "freeze_header": True,
            "auto_filter": True,
            "auto_width": True,
        },
    }


class TransformacionExcelApiIntegrationTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.engine = create_test_engine()
        from fastapi.testclient import TestClient

        cls.test_client_class = TestClient

    @classmethod
    def tearDownClass(cls) -> None:
        cls.engine.dispose()

    def setUp(self) -> None:
        self.isolation = IsolatedDatabaseAndStorage(self.engine)
        self.isolation.start()

        def override_get_db():
            session = self.isolation.session()
            try:
                yield session
            finally:
                session.close()

        app.dependency_overrides[get_db] = override_get_db
        self.client = self.test_client_class(app)
        self._seed_database()
        self.admin_headers = self._login("admin@example.test")
        self.user_headers = self._login("user@example.test")
        self.other_headers = self._login("other@example.test")

    def tearDown(self) -> None:
        self.client.close()
        app.dependency_overrides.clear()
        self.isolation.stop()

    def _seed_database(self) -> None:
        with self.isolation.session() as db:
            primary = Cliente(nombre="Cliente principal", cuit="20111111112")
            other = Cliente(nombre="Otro cliente", cuit="20222222223")
            db.add_all([primary, other])
            db.flush()
            admin = Usuario(
                cliente_id=primary.id,
                nombre="Admin",
                email="admin@example.test",
                password_hash=get_password_hash(PASSWORD),
                rol="ADMIN",
                estado="ACTIVO",
            )
            user = Usuario(
                cliente_id=primary.id,
                nombre="Operador",
                email="user@example.test",
                password_hash=get_password_hash(PASSWORD),
                rol="OPERADOR",
                estado="ACTIVO",
            )
            other_user = Usuario(
                cliente_id=other.id,
                nombre="Otro",
                email="other@example.test",
                password_hash=get_password_hash(PASSWORD),
                rol="ADMIN",
                estado="ACTIVO",
            )
            process = Proceso(
                cliente_id=primary.id,
                nombre="Transformar ventas",
                tipo="TRANSFORMACION_EXCEL",
            )
            secondary_process = Proceso(
                cliente_id=primary.id,
                nombre="Otra transformación",
                tipo="TRANSFORMACION_EXCEL",
            )
            other_process = Proceso(
                cliente_id=other.id,
                nombre="Transformar otro",
                tipo="TRANSFORMACION_EXCEL",
            )
            wrong_process = Proceso(
                cliente_id=primary.id,
                nombre="Conciliar",
                tipo="CONCILIACION",
            )
            db.add_all(
                [
                    admin,
                    user,
                    other_user,
                    process,
                    secondary_process,
                    other_process,
                    wrong_process,
                ],
            )
            db.commit()
            self.admin_id = admin.id
            self.user_id = user.id
            self.process_id = process.id
            self.secondary_process_id = secondary_process.id
            self.other_process_id = other_process.id
            self.wrong_process_id = wrong_process.id

    def _login(self, email: str) -> dict[str, str]:
        response = self.client.post(
            "/auth/login",
            json={"email": email, "password": PASSWORD},
        )
        self.assertEqual(response.status_code, 200, response.text)
        return {"Authorization": f"Bearer {response.json()['access_token']}"}

    def _create_execution(self, process_id: int | None = None) -> int:
        response = self.client.post(
            "/ejecuciones",
            json={"proceso_id": process_id or self.process_id},
            headers=self.user_headers,
        )
        self.assertEqual(response.status_code, 201, response.text)
        return response.json()["id"]

    def _upload(
        self,
        execution_id: int,
        content: bytes | None = None,
        filename: str = "ventas.xlsx",
    ) -> int:
        response = self.client.post(
            "/archivos/upload",
            data={"ejecucion_id": execution_id, "tipo_archivo": "FUENTE"},
            files={
                "file": (
                    filename,
                    content if content is not None else workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
            headers=self.user_headers,
        )
        self.assertEqual(response.status_code, 201, response.text)
        return response.json()["id"]

    def _configure_and_validate(self, execution_id: int, archivo_id: int) -> dict:
        config = full_config(archivo_id)
        response = self.client.post(
            f"/transformaciones-excel/{execution_id}/configuracion",
            json=config,
            headers=self.user_headers,
        )
        self.assertEqual(response.status_code, 200, response.text)
        response = self.client.post(
            f"/transformaciones-excel/{execution_id}/validar",
            headers=self.user_headers,
        )
        self.assertEqual(response.status_code, 200, response.text)
        self.assertTrue(response.json()["valid"])
        return config

    def _physical_path(self, archivo_id: int) -> Path:
        with self.isolation.session() as db:
            record = db.get(Archivo, archivo_id)
            return resolve_storage_path_safely(
                record.ruta_storage,
                self.isolation.storage_root,
            )

    def test_openapi_and_compatibility_routes(self) -> None:
        self.assertEqual(self.client.get("/health").status_code, 200)
        token_response = self.client.post(
            "/auth/token",
            data={"username": "user@example.test", "password": PASSWORD},
        )
        self.assertEqual(token_response.status_code, 200, token_response.text)
        schema_response = self.client.get("/openapi.json")
        self.assertEqual(schema_response.status_code, 200)
        schema = schema_response.json()
        required_paths = {
            "/transformaciones-excel/archivos/{archivo_id}/estructura",
            "/transformaciones-excel/{ejecucion_id}/configuracion",
            "/transformaciones-excel/{ejecucion_id}/validar",
            "/transformaciones-excel/{ejecucion_id}/generar",
            "/transformaciones-excel/{ejecucion_id}/resultado",
            "/transformaciones-excel/{ejecucion_id}/resultado/descargar",
            "/transformaciones-excel/{ejecucion_id}/resumen",
            "/transformaciones-excel/{ejecucion_id}/trazabilidad",
            "/transformaciones-excel/procesos/{proceso_id}/plantillas",
            "/transformaciones-excel/plantillas/{plantilla_id}",
            "/transformaciones-excel/{ejecucion_id}/plantillas",
            (
                "/transformaciones-excel/{ejecucion_id}/plantillas/"
                "{plantilla_id}/aplicar"
            ),
        }
        self.assertTrue(required_paths.issubset(schema["paths"]))
        self.assertTrue(any(path.startswith("/conciliaciones") for path in schema["paths"]))
        operation_ids = [
            operation["operationId"]
            for path in schema["paths"].values()
            for operation in path.values()
            if isinstance(operation, dict) and "operationId" in operation
        ]
        self.assertEqual(len(operation_ids), len(set(operation_ids)))

    def test_authentication_authorization_and_process_type(self) -> None:
        execution_id = self._create_execution()
        archivo_id = self._upload(execution_id)
        url = f"/transformaciones-excel/archivos/{archivo_id}/estructura"
        self.assertEqual(self.client.get(url).status_code, 401)
        self.assertEqual(
            self.client.get(url, headers=self.other_headers).status_code,
            403,
        )

        wrong_execution = self._create_execution(self.wrong_process_id)
        wrong_file = self._upload(wrong_execution)
        response = self.client.post(
            f"/transformaciones-excel/{wrong_execution}/configuracion",
            json=full_config(wrong_file),
            headers=self.user_headers,
        )
        self.assertEqual(response.status_code, 400)

        self._configure_and_validate(execution_id, archivo_id)
        response = self.client.post(
            f"/transformaciones-excel/{execution_id}/plantillas",
            json={"nombre": "Base"},
            headers=self.user_headers,
        )
        self.assertEqual(response.status_code, 403)
        response = self.client.post(
            f"/transformaciones-excel/{execution_id}/plantillas",
            json={"nombre": "Base"},
            headers=self.admin_headers,
        )
        self.assertEqual(response.status_code, 201, response.text)
        template_id = response.json()["id"]
        self.assertEqual(
            self.client.put(
                f"/transformaciones-excel/plantillas/{template_id}",
                json={"nombre": "Sin permiso"},
                headers=self.user_headers,
            ).status_code,
            403,
        )
        apply_execution = self._create_execution()
        apply_file = self._upload(apply_execution)
        response = self.client.post(
            (
                f"/transformaciones-excel/{apply_execution}/plantillas/"
                f"{template_id}/aplicar"
            ),
            json={"archivo_id": apply_file},
            headers=self.user_headers,
        )
        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(
            self.client.delete(
                f"/transformaciones-excel/plantillas/{template_id}",
                headers=self.user_headers,
            ).status_code,
            403,
        )
        self.assertEqual(
            self.client.delete(
                f"/transformaciones-excel/plantillas/{template_id}",
                headers=self.admin_headers,
            ).status_code,
            204,
        )
        response = self.client.post(
            (
                f"/transformaciones-excel/{execution_id}/plantillas/"
                f"{template_id}/aplicar"
            ),
            json={"archivo_id": archivo_id},
            headers=self.user_headers,
        )
        self.assertEqual(response.status_code, 400)

        secondary_execution = self._create_execution(self.secondary_process_id)
        secondary_file = self._upload(secondary_execution)
        self._configure_and_validate(secondary_execution, secondary_file)
        response = self.client.post(
            f"/transformaciones-excel/{secondary_execution}/plantillas",
            json={"nombre": "Otro proceso"},
            headers=self.admin_headers,
        )
        self.assertEqual(response.status_code, 201, response.text)
        other_template_id = response.json()["id"]
        response = self.client.post(
            (
                f"/transformaciones-excel/{execution_id}/plantillas/"
                f"{other_template_id}/aplicar"
            ),
            json={"archivo_id": archivo_id},
            headers=self.user_headers,
        )
        self.assertEqual(response.status_code, 400)

    def test_complete_flow_generation_reuse_summary_trace_and_download(self) -> None:
        execution_id = self._create_execution()
        archivo_id = self._upload(execution_id)
        inspection = self.client.get(
            f"/transformaciones-excel/archivos/{archivo_id}/estructura",
            headers=self.user_headers,
        )
        self.assertEqual(inspection.status_code, 200, inspection.text)
        self.assertEqual(inspection.json()["available_sheets"], ["Ventas"])
        self.assertEqual(
            [column["name"] for column in inspection.json()["columns"]],
            ["Cliente", "Cantidad", "Precio", "Categoria"],
        )

        config = self._configure_and_validate(execution_id, archivo_id)
        saved = self.client.get(
            f"/transformaciones-excel/{execution_id}/configuracion",
            headers=self.user_headers,
        )
        self.assertEqual(saved.status_code, 200)
        self.assertEqual(saved.json()["configuracion"], config)

        generated = self.client.post(
            f"/transformaciones-excel/{execution_id}/generar",
            headers=self.user_headers,
        )
        self.assertEqual(generated.status_code, 200, generated.text)
        self.assertEqual(generated.json()["estado_ejecucion"], "COMPLETADO")
        self.assertFalse(generated.json()["reused"])
        second = self.client.post(
            f"/transformaciones-excel/{execution_id}/generar",
            headers=self.user_headers,
        )
        self.assertEqual(second.status_code, 200, second.text)
        self.assertTrue(second.json()["reused"])

        result = self.client.get(
            f"/transformaciones-excel/{execution_id}/resultado",
            headers=self.user_headers,
        )
        self.assertEqual(result.status_code, 200)
        download = self.client.get(
            f"/transformaciones-excel/{execution_id}/resultado/descargar",
            headers=self.user_headers,
        )
        self.assertEqual(download.status_code, 200, download.text)
        self.assertEqual(download.headers["x-content-type-options"], "nosniff")
        self.assertEqual(download.headers["cache-control"], "private, no-store")
        self.assertNotIn("ruta_storage", download.text)
        workbook = load_workbook(BytesIO(download.content), data_only=False)
        worksheet = workbook["Resultado"]
        self.assertEqual(
            [cell.value for cell in worksheet[1]],
            ["Cliente", "Empresa", "Etiqueta", "Total", "Categoria normalizada"],
        )
        self.assertEqual(worksheet.max_row, 3)

        summary = self.client.get(
            f"/transformaciones-excel/{execution_id}/resumen",
            headers=self.user_headers,
        ).json()
        self.assertEqual(summary["action_required"], "DOWNLOAD")
        trace = self.client.get(
            f"/transformaciones-excel/{execution_id}/trazabilidad",
            headers=self.user_headers,
        ).json()
        event_types = {event["event_type"] for event in trace["items"]}
        self.assertTrue(
            {
                "CONFIGURATION_SAVED",
                "VALIDATION_SUCCEEDED",
                "GENERATION_STARTED",
                "GENERATION_COMPLETED",
                "GENERATION_REUSED",
            }.issubset(event_types),
        )
        with self.isolation.session() as db:
            count = db.scalar(
                select(func.count(Archivo.id)).where(
                    Archivo.ejecucion_id == execution_id,
                    Archivo.tipo_archivo == OUTPUT_FILE_TYPE,
                ),
            )
        self.assertEqual(count, 1)

    def test_functional_validation_errors_and_generation_preconditions(self) -> None:
        execution_id = self._create_execution()
        archivo_id = self._upload(execution_id)
        missing_column = full_config(archivo_id)
        missing_column["output_columns"][0]["source_column"] = "No existe"
        response = self.client.post(
            f"/transformaciones-excel/{execution_id}/configuracion",
            json=missing_column,
            headers=self.user_headers,
        )
        self.assertEqual(response.status_code, 400)

        other_execution = self._create_execution()
        response = self.client.post(
            f"/transformaciones-excel/{other_execution}/configuracion",
            json=full_config(archivo_id),
            headers=self.user_headers,
        )
        self.assertEqual(response.status_code, 400)

        response = self.client.post(
            f"/transformaciones-excel/{execution_id}/configuracion",
            json=full_config(archivo_id),
            headers=self.user_headers,
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            self.client.post(
                f"/transformaciones-excel/{execution_id}/generar",
                headers=self.user_headers,
            ).status_code,
            409,
        )

        unmapped_file = self._upload(
            other_execution,
            workbook_bytes([["Ana", 2, 10, "SIN_MAPEO"]]),
        )
        config = full_config(unmapped_file)
        self.assertEqual(
            self.client.post(
                f"/transformaciones-excel/{other_execution}/configuracion",
                json=config,
                headers=self.user_headers,
            ).status_code,
            200,
        )
        validation = self.client.post(
            f"/transformaciones-excel/{other_execution}/validar",
            headers=self.user_headers,
        )
        self.assertEqual(validation.status_code, 200)
        self.assertFalse(validation.json()["valid"])

        division_execution = self._create_execution()
        division_file = self._upload(
            division_execution,
            workbook_bytes([["Ana", 2, 0, "A"]]),
        )
        division_config = full_config(division_file)
        arithmetic = division_config["output_columns"][3]
        arithmetic["operator"] = "DIVIDE"
        self.assertEqual(
            self.client.post(
                f"/transformaciones-excel/{division_execution}/configuracion",
                json=division_config,
                headers=self.user_headers,
            ).status_code,
            200,
        )
        validation = self.client.post(
            f"/transformaciones-excel/{division_execution}/validar",
            headers=self.user_headers,
        )
        self.assertEqual(validation.status_code, 200)
        self.assertFalse(validation.json()["valid"])
        self.assertEqual(
            self.client.post(
                f"/transformaciones-excel/{division_execution}/generar",
                headers=self.user_headers,
            ).status_code,
            409,
        )

    def test_source_config_and_historical_integrity_invalidation(self) -> None:
        execution_id = self._create_execution()
        archivo_id = self._upload(execution_id)
        self._configure_and_validate(execution_id, archivo_id)
        self._physical_path(archivo_id).write_bytes(
            workbook_bytes([["Ana", 2, 11, "A"]]),
        )
        response = self.client.post(
            f"/transformaciones-excel/{execution_id}/generar",
            headers=self.user_headers,
        )
        self.assertEqual(response.status_code, 409, response.text)
        self.assertIn("SOURCE_CHANGED_AFTER_VALIDATION", response.text)

        execution_id = self._create_execution()
        archivo_id = self._upload(execution_id)
        config = self._configure_and_validate(execution_id, archivo_id)
        config["output"]["sheet_name"] = "Configuracion nueva"
        self.assertEqual(
            self.client.post(
                f"/transformaciones-excel/{execution_id}/configuracion",
                json=config,
                headers=self.user_headers,
            ).status_code,
            200,
        )
        response = self.client.post(
            f"/transformaciones-excel/{execution_id}/generar",
            headers=self.user_headers,
        )
        self.assertEqual(response.status_code, 409, response.text)
        self.assertIn("CONFIG_CHANGED_AFTER_VALIDATION", response.text)

        execution_id = self._create_execution()
        archivo_id = self._upload(execution_id)
        self._configure_and_validate(execution_id, archivo_id)
        with self.isolation.session() as db:
            execution = db.get(EjecucionProceso, execution_id)
            summary = deepcopy(execution.resumen_json)
            summary["transformacion_excel"]["configuracion"]["output"][
                "sheet_name"
            ] = "Cambiada"
            execution.resumen_json = summary
            db.commit()
        response = self.client.post(
            f"/transformaciones-excel/{execution_id}/generar",
            headers=self.user_headers,
        )
        self.assertEqual(response.status_code, 409, response.text)
        self.assertIn("CONFIG_CHANGED_AFTER_VALIDATION", response.text)

        execution_id = self._create_execution()
        archivo_id = self._upload(execution_id)
        self._configure_and_validate(execution_id, archivo_id)
        with self.isolation.session() as db:
            execution = db.get(EjecucionProceso, execution_id)
            summary = deepcopy(execution.resumen_json)
            validation = summary["transformacion_excel"]["validacion"]
            validation.pop("source_checksum")
            validation.pop("config_checksum")
            execution.resumen_json = summary
            db.commit()
        response = self.client.post(
            f"/transformaciones-excel/{execution_id}/generar",
            headers=self.user_headers,
        )
        self.assertEqual(response.status_code, 409, response.text)
        self.assertIn("VALIDATION_INTEGRITY_DATA_MISSING", response.text)

    def test_security_limits_corrupt_zip_unsafe_path_and_output_name(self) -> None:
        execution_id = self._create_execution()
        archivo_id = self._upload(execution_id)
        with patch.object(settings, "transformacion_excel_max_rows", 1):
            response = self.client.get(
                f"/transformaciones-excel/archivos/{archivo_id}/estructura",
                headers=self.user_headers,
            )
        self.assertEqual(response.status_code, 413)
        self.assertIn("SOURCE_ROW_LIMIT_EXCEEDED", response.text)

        corrupt_id = self._upload(execution_id, b"not-a-zip", "corrupt.xlsx")
        response = self.client.get(
            f"/transformaciones-excel/archivos/{corrupt_id}/estructura",
            headers=self.user_headers,
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("INVALID_XLSX_CONTAINER", response.text)

        unsafe = BytesIO()
        with ZipFile(unsafe, "w") as archive:
            archive.writestr("../escape", "unsafe")
        unsafe_id = self._upload(execution_id, unsafe.getvalue(), "unsafe.xlsx")
        response = self.client.get(
            f"/transformaciones-excel/archivos/{unsafe_id}/estructura",
            headers=self.user_headers,
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("UNSAFE_XLSX_ENTRY", response.text)

        with self.isolation.session() as db:
            record = db.get(Archivo, archivo_id)
            record.ruta_storage = "../outside.xlsx"
            db.commit()
        response = self.client.get(
            f"/transformaciones-excel/archivos/{archivo_id}/estructura",
            headers=self.user_headers,
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("UNSAFE_STORAGE_PATH", response.text)
        self.assertNotIn(str(self.isolation.storage_root), response.text)

    def test_file_column_and_xlsx_expansion_limits_return_413(self) -> None:
        execution_id = self._create_execution()
        large_file = self._upload(
            execution_id,
            b"value\n" + b"x" * (1024 * 1024),
            "large.csv",
        )
        with patch.object(settings, "transformacion_excel_max_file_size_mb", 1):
            response = self.client.get(
                f"/transformaciones-excel/archivos/{large_file}/estructura",
                headers=self.user_headers,
            )
        self.assertEqual(response.status_code, 413)
        self.assertIn("SOURCE_FILE_TOO_LARGE", response.text)

        archivo_id = self._upload(execution_id)
        with patch.object(settings, "transformacion_excel_max_columns", 3):
            response = self.client.get(
                f"/transformaciones-excel/archivos/{archivo_id}/estructura",
                headers=self.user_headers,
            )
        self.assertEqual(response.status_code, 413)
        self.assertIn("SOURCE_COLUMN_LIMIT_EXCEEDED", response.text)

        compressed = BytesIO()
        workbook_xml = (
            '<workbook xmlns="http://schemas.openxmlformats.org/'
            'spreadsheetml/2006/main"><sheets/></workbook>'
        )
        with ZipFile(compressed, "w", compression=ZIP_DEFLATED) as archive:
            archive.writestr("xl/workbook.xml", workbook_xml)
            archive.writestr("xl/worksheets/sheet1.xml", b"0" * 1_100_000)
        compressed_id = self._upload(
            execution_id,
            compressed.getvalue(),
            "compressed.xlsx",
        )
        with patch.object(
            settings,
            "transformacion_excel_max_xlsx_compression_ratio",
            2,
        ):
            response = self.client.get(
                f"/transformaciones-excel/archivos/{compressed_id}/estructura",
                headers=self.user_headers,
            )
        self.assertEqual(response.status_code, 413)
        self.assertIn("XLSX_COMPRESSION_RATIO_LIMIT", response.text)

    def test_formula_injection_is_neutralized_and_output_traversal_is_rejected(self) -> None:
        execution_id = self._create_execution()
        csv_content = (
            b"Cliente,Cantidad,Precio,Categoria\n"
            b"=2+2,1,2,A\n"
        )
        archivo_id = self._upload(execution_id, csv_content, "ventas.csv")
        config = full_config(archivo_id)
        config["source"]["sheet_name"] = None
        self.assertEqual(
            self.client.post(
                f"/transformaciones-excel/{execution_id}/configuracion",
                json=config,
                headers=self.user_headers,
            ).status_code,
            200,
        )
        self.assertTrue(
            self.client.post(
                f"/transformaciones-excel/{execution_id}/validar",
                headers=self.user_headers,
            ).json()["valid"],
        )
        self.assertEqual(
            self.client.post(
                f"/transformaciones-excel/{execution_id}/generar",
                headers=self.user_headers,
            ).status_code,
            200,
        )
        download = self.client.get(
            f"/transformaciones-excel/{execution_id}/resultado/descargar",
            headers=self.user_headers,
        )
        workbook = load_workbook(BytesIO(download.content), data_only=False)
        self.assertEqual(workbook["Resultado"]["A2"].value, "'=2+2")

        traversal_execution = self._create_execution()
        traversal_file = self._upload(traversal_execution)
        traversal_config = full_config(traversal_file, "../escape.xlsx")
        self.assertEqual(
            self.client.post(
                f"/transformaciones-excel/{traversal_execution}/configuracion",
                json=traversal_config,
                headers=self.user_headers,
            ).status_code,
            200,
        )
        self.assertTrue(
            self.client.post(
                f"/transformaciones-excel/{traversal_execution}/validar",
                headers=self.user_headers,
            ).json()["valid"],
        )
        response = self.client.post(
            f"/transformaciones-excel/{traversal_execution}/generar",
            headers=self.user_headers,
        )
        self.assertEqual(response.status_code, 400)
        self.assertNotIn(str(self.isolation.storage_root), response.text)

    def test_missing_source_and_output_are_reported_operationally(self) -> None:
        execution_id = self._create_execution()
        source_id = self._upload(execution_id)
        self._configure_and_validate(execution_id, source_id)
        generated = self.client.post(
            f"/transformaciones-excel/{execution_id}/generar",
            headers=self.user_headers,
        )
        self.assertEqual(generated.status_code, 200, generated.text)
        output_id = generated.json()["archivo_id"]
        self._physical_path(output_id).unlink()
        summary = self.client.get(
            f"/transformaciones-excel/{execution_id}/resumen",
            headers=self.user_headers,
        ).json()
        self.assertEqual(summary["action_required"], "REGENERATE")
        self.assertIn("OUTPUT_FILE_MISSING", [item["code"] for item in summary["issues"]])

        self._physical_path(source_id).unlink()
        summary = self.client.get(
            f"/transformaciones-excel/{execution_id}/resumen",
            headers=self.user_headers,
        ).json()
        source_issue = next(
            item for item in summary["issues"] if item["code"] == "SOURCE_FILE_MISSING"
        )
        self.assertTrue(source_issue["blocking"])


if __name__ == "__main__":
    unittest.main()
