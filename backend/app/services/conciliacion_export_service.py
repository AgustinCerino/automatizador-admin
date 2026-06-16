from datetime import date, datetime
from decimal import Decimal
import json
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.models import ResultadoConciliacion
from app.services.conciliacion_revision_service import (
    ConciliacionRevisionError,
    build_revision_summary,
    get_ejecucion,
    get_resultados_ejecucion,
)


PROCESSED_STORAGE_ROOT = Path(__file__).resolve().parents[2] / "storage" / "processed"

RESULT_COLUMNS = [
    "id",
    "ejecucion_id",
    "clave_referencia",
    "estado_resultado",
    "diferencia_importe",
    "requiere_revision",
    "observacion",
    "datos_archivo_a_json",
    "datos_archivo_b_json",
    "created_at",
    "updated_at",
]


def excel_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime | date):
        return value.isoformat()
    if isinstance(value, dict | list):
        return json.dumps(value, ensure_ascii=False, indent=2, default=str)
    return value


def resultado_to_row(resultado: ResultadoConciliacion) -> list[Any]:
    return [
        excel_value(resultado.id),
        excel_value(resultado.ejecucion_id),
        excel_value(resultado.clave_referencia),
        excel_value(resultado.estado_resultado),
        excel_value(resultado.diferencia_importe),
        excel_value(resultado.requiere_revision),
        excel_value(resultado.observacion),
        excel_value(resultado.datos_archivo_a_json),
        excel_value(resultado.datos_archivo_b_json),
        excel_value(resultado.created_at),
        excel_value(resultado.updated_at),
    ]


def apply_table_format(sheet: Worksheet) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    sheet.freeze_panes = "A2"

    if sheet.max_column > 1:
        sheet.auto_filter.ref = sheet.dimensions

    for column_cells in sheet.columns:
        column_letter = column_cells[0].column_letter
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        sheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 12),
            60,
        )


def append_result_sheet(
    workbook: Workbook,
    title: str,
    resultados: Iterable[ResultadoConciliacion],
) -> None:
    sheet = workbook.create_sheet(title=title)
    sheet.append(RESULT_COLUMNS)
    for resultado in resultados:
        sheet.append(resultado_to_row(resultado))
    apply_table_format(sheet)


def write_summary_sheet(
    sheet: Worksheet,
    summary: dict[str, Any],
) -> None:
    sheet.title = "Resumen"
    sheet.append(["campo", "valor"])
    for key, value in summary.items():
        sheet.append([key, excel_value(value)])
    apply_table_format(sheet)


def ensure_processed_dir(ejecucion_id: int) -> Path:
    directory = PROCESSED_STORAGE_ROOT / str(ejecucion_id)
    directory.mkdir(parents=True, exist_ok=True)
    return directory


def export_reconciliation_results(db: Session, ejecucion_id: int) -> Path:
    ejecucion = get_ejecucion(db, ejecucion_id)
    resultados = get_resultados_ejecucion(db, ejecucion_id)

    if not resultados:
        raise ConciliacionRevisionError(
            "La ejecución no tiene resultados para exportar",
        )

    summary = build_revision_summary(ejecucion, resultados)
    workbook = Workbook()
    write_summary_sheet(workbook.active, summary)

    append_result_sheet(workbook, "Todos", resultados)
    append_result_sheet(
        workbook,
        "Conciliados",
        [
            resultado
            for resultado in resultados
            if resultado.estado_resultado == "CONCILIADO"
        ],
    )
    append_result_sheet(
        workbook,
        "Diferencias de importe",
        [
            resultado
            for resultado in resultados
            if resultado.estado_resultado == "DIFERENCIA_IMPORTE"
        ],
    )
    append_result_sheet(
        workbook,
        "Solo Archivo A",
        [
            resultado
            for resultado in resultados
            if resultado.estado_resultado == "SOLO_ARCHIVO_A"
        ],
    )
    append_result_sheet(
        workbook,
        "Solo Archivo B",
        [
            resultado
            for resultado in resultados
            if resultado.estado_resultado == "SOLO_ARCHIVO_B"
        ],
    )
    append_result_sheet(
        workbook,
        "Duplicados",
        [
            resultado
            for resultado in resultados
            if resultado.estado_resultado
            in {"DUPLICADO_ARCHIVO_A", "DUPLICADO_ARCHIVO_B"}
        ],
    )
    append_result_sheet(
        workbook,
        "Errores de formato",
        [
            resultado
            for resultado in resultados
            if resultado.estado_resultado == "ERROR_FORMATO"
        ],
    )
    append_result_sheet(
        workbook,
        "Revisión manual",
        [
            resultado
            for resultado in resultados
            if resultado.requiere_revision or resultado.observacion is not None
        ],
    )

    output_dir = ensure_processed_dir(ejecucion_id)
    output_path = output_dir / f"conciliacion_ejecucion_{ejecucion_id}.xlsx"
    workbook.save(output_path)
    return output_path
